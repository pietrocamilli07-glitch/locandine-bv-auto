import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

args = sys.argv[1:]
ins = [a for a in args if a.endswith('.json')] or ['players_all.json']
outs = [a for a in args if a.endswith('.xlsx')]
OUT = outs[0] if outs else 'Rubrica-Giocatori-2026.xlsx'
players = []
for f in ins:
    players += json.load(open(f, encoding='utf-8'))

GEN_ORDER = ['MASCHILE', 'FEMMINILE']
CAT_ORDER = ['UNDER 14', 'UNDER 16', 'UNDER 18', 'UNDER 20', 'ASSOLUTO']
CAT_SHORT = {'UNDER 14': 'U14', 'UNDER 16': 'U16', 'UNDER 18': 'U18', 'UNDER 20': 'U20', 'ASSOLUTO': 'Assoluto'}
GEN_SHORT = {'MASCHILE': 'Maschile', 'FEMMINILE': 'Femminile'}
GEN_COL = {'MASCHILE': '1F4E79', 'FEMMINILE': 'B0286B'}

groups = {}
for p in players:
    groups.setdefault((p['genere'], p['categoria']), {})[p['id']] = p['name']

def tc(n):
    return ' '.join(w.capitalize() for w in n.split())

FNT = 'Arial'
title_font = Font(name=FNT, bold=True, size=13, color='FFFFFF')
note_font = Font(name=FNT, italic=True, size=9, color='FFFFFF')
hdr_font = Font(name=FNT, bold=True, size=10, color='FFFFFF')
id_font = Font(name=FNT, size=9, color='9AA0A6')
name_font = Font(name=FNT, size=11, color='000000')
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill('solid', fgColor='404040')
id_fill = PatternFill('solid', fgColor='F2F2F2')
at_fill = PatternFill('solid', fgColor='FFF8E1')
UNLOCK = Protection(locked=False)

wb = Workbook()
wb.remove(wb.active)
tot = 0
for g in GEN_ORDER:
    for cat in CAT_ORDER:
        d = groups.get((g, cat))
        if not d:
            continue
        rws = sorted(d.items(), key=lambda kv: (kv[1], kv[0]))
        ws = wb.create_sheet(title=GEN_SHORT[g] + ' ' + CAT_SHORT[cat])
        ws.sheet_properties.tabColor = GEN_COL[g]
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 34
        ws.column_dimensions['C'].width = 28
        ws.merge_cells('A1:C1')
        c = ws.cell(1, 1, g + '  -  ' + cat + '   (' + str(len(rws)) + ' giocatori)')
        c.font = title_font; c.fill = PatternFill('solid', fgColor=GEN_COL[g])
        c.alignment = Alignment(horizontal='left', vertical='center'); ws.row_dimensions[1].height = 24
        ws.merge_cells('A2:C2')
        c = ws.cell(2, 1, 'Scrivi SOLO la tua @ Instagram nella colonna C (gialla). ID e Cognome/Nome non sono modificabili.')
        c.font = note_font; c.fill = PatternFill('solid', fgColor=GEN_COL[g])
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True); ws.row_dimensions[2].height = 22
        for col, txt in enumerate(['ID federvolley (uso interno)', 'COGNOME NOME', '@ Instagram (da compilare)'], 1):
            hc = ws.cell(3, col, txt); hc.font = hdr_font; hc.fill = hdr_fill
            hc.alignment = Alignment(horizontal='left', vertical='center'); hc.border = border
        r = 4
        for pid, name in rws:
            a = ws.cell(r, 1, pid); a.font = id_font; a.fill = id_fill; a.border = border
            a.alignment = Alignment(horizontal='center')
            b = ws.cell(r, 2, tc(name)); b.font = name_font; b.border = border
            cc = ws.cell(r, 3, ''); cc.fill = at_fill; cc.border = border; cc.protection = UNLOCK
            r += 1
            tot += 1
        ws.freeze_panes = 'A4'
        ws.protection.sheet = True
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.sort = True
        ws.protection.autoFilter = True
wb.save(OUT)
print('Fogli:', [s.title for s in wb.worksheets])
print('Totale righe giocatore:', tot)
